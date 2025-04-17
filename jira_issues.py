from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
import os
from dotenv import load_dotenv, dotenv_values
from datetime import datetime, timezone
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from dateutil import parser
from urllib.parse import quote
import win32com.client
from pathlib import Path


def create_chrome_driver(headless=False):
    """
    Creates and returns a configured Chrome WebDriver instance

    Args:
        headless (bool): Whether to run Chrome in headless mode

    Returns:
        WebDriver: Configured Chrome webdriver instance
    """
    print("Setting up Chrome options...")
    options = Options()
    options.add_argument("--start-maximized")

    if headless:
        options.add_argument("--headless")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920,1080")

    print("Initializing Chrome browser...")
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=options
    )

    return driver


def navigate_to_url(driver, url, wait_element=None, timeout=10):
    """
    Navigates to a specified URL using the provided WebDriver and waits for an element to load

    Args:
        driver (WebDriver): Chrome WebDriver instance
        url (str): URL to navigate to
        wait_element (str, optional): CSS selector or ID of element to wait for
        timeout (int): Maximum time to wait for the element in seconds

    Returns:
        str: The title of the loaded page
    """
    print(f"Navigating to: {url}")
    driver.get(url)

    if wait_element:
        # Determine if the selector is an ID or a class
        if wait_element.startswith("#"):
            print(f"Waiting for element with ID: {wait_element[1:]}")
            by_method = By.ID
            selector = wait_element[1:]
        elif wait_element.startswith("."):
            print(f"Waiting for element with class: {wait_element[1:]}")
            by_method = By.CLASS_NAME
            selector = wait_element[1:]
        else:
            print(f"Waiting for element with selector: {wait_element}")
            by_method = By.CSS_SELECTOR
            selector = wait_element

        try:
            print(f"Waiting for element to load (timeout: {timeout}s)...")
            element = WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((by_method, selector))
            )
            print(f"Element found: {element.tag_name}")
        except Exception as e:
            print(f"Timeout waiting for element: {wait_element}")
            print(f"Error: {e}")
    else:
        # Default wait for page to load if no specific element is specified
        print("No specific element to wait for. Waiting for page to load...")
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )

    page_title = driver.title
    print(f"Current page title: {page_title}")

    return page_title


def load_jira_filters(filter_pattern=None):
    """
    Loads all environment variables defined in the .env file that begin with the specified pattern.

    Args:
        filter_pattern (str): The pattern to use for filtering environment variables

    Returns:
        dict: A dictionary where keys are the environment variable names and values are their contents.
    """
    # Load all variables from the .env file
    env_vars = dotenv_values()

    # Filter variables that start with the specified pattern
    filters = {
        key: value for key, value in env_vars.items() if key.startswith(filter_pattern)
    }

    print(f"Found {len(filters)} JIRA filters matching pattern '{filter_pattern}'")

    return filters


def load_mail_recipients(pattern_type="TO"):
    """
    Loads email addresses defined in the .env file that begin with the specified pattern.

    Args:
        pattern_type (str): Type of recipients to load ("TO" or "CC")

    Returns:
        str: String containing all email addresses separated by semicolons.
    """
    # Determine the pattern based on recipient type
    if pattern_type == "CC":
        mail_pattern = os.getenv("MAIL_PATTERN_CC", "MAIL_CC_")
    else:
        mail_pattern = os.getenv("MAIL_PATTERN", "MAIL_TO_")

    # Load all variables from the .env file
    env_vars = dotenv_values()

    # Filter variables that start with the specified pattern
    mail_vars = {
        key: value for key, value in env_vars.items() if key.startswith(mail_pattern)
    }

    # Extract only the values (email addresses) and join them with semicolons
    recipients = ";".join(mail_vars.values())

    print(
        f"Found {len(mail_vars)} email addresses for {pattern_type} list matching pattern '{mail_pattern}'"
    )

    return recipients


def extract_jira_issues(driver):
    """
    Extracts JIRA issue information from the issue table and stores it in a pandas DataFrame.

    This function finds all rows in the JIRA issue table and extracts the following data for each issue:
    - Issue Key (e.g., PROJECT-123)
    - Issue Type (e.g., Bug, Story, Task)
    - Issue Link (URL to the issue)
    - Summary (issue title/description)
    - Status (e.g., To Do, In Progress, Done)
    - Priority (e.g., Highest, High, Medium, Low)
    - Customer Object ID (custom field)
    - Assignee (person assigned to the issue)
    - Created Date (when the issue was created)
    - Classification (custom field)

    Args:
        driver (WebDriver): Chrome WebDriver instance with JIRA page already loaded

    Returns:
        DataFrame: Pandas DataFrame containing all extracted issue information
    """
    print("Extracting JIRA issues from the table...")

    # Wait for the issue table to be present in the DOM
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "issuetable"))
    )

    # Find all rows in the table body
    rows = driver.find_elements(By.CSS_SELECTOR, "#issuetable tbody tr")

    # Initialize lists to store issue data
    issue_keys = []
    issue_types = []
    issue_links = []
    summaries = []
    statuses = []
    priorities = []
    customer_object_ids = []
    assignees = []
    created_dates = []
    classifications = []

    # Extract data from each row in the issue table
    for row in rows:
        # Get the issue key from the row attribute
        issue_key = row.get_attribute("data-issuekey")

        # Get the issue link
        issue_link = ""
        try:
            # Find the td with class "issuekey"
            issuekey_td = row.find_element(By.CSS_SELECTOR, "td.issuekey")

            # Find the a with class "issue-link" inside the td
            issue_link_element = issuekey_td.find_element(
                By.CSS_SELECTOR, "a.issue-link"
            )

            # Get the href attribute
            issue_link = issue_link_element.get_attribute("href")
        except Exception as e:
            print(f"Error getting link for issue {issue_key}: {e}")

        # Get the type of issue
        issue_type = ""
        try:
            # Find the td with class "issuetype"
            issue_type_td = row.find_element(By.CSS_SELECTOR, "td.issuetype")

            # Find the img element inside the issuetype td
            issue_type_img = issue_type_td.find_element(By.CSS_SELECTOR, "img")

            # Get the alt attribute which contains the issue type
            issue_type = issue_type_img.get_attribute("alt").strip()
        except Exception as e:
            print(f"Error getting issue type for issue {issue_key}: {e}")

        # Get the summary
        summary = ""
        try:
            # Find the td with class "summary"
            summary_td = row.find_element(By.CSS_SELECTOR, "td.summary")

            # Find the p element inside the summary td
            summary_element = summary_td.find_element(By.CSS_SELECTOR, "p")

            # Get the text content
            summary = summary_element.text.strip()
        except Exception as e:
            print(f"Error getting summary for issue {issue_key}: {e}")

        # Get the status
        status = ""
        try:
            # Find the td with class "status"
            status_td = row.find_element(By.CSS_SELECTOR, "td.status")

            # Find the span element inside the status td
            status_element = status_td.find_element(By.CSS_SELECTOR, "span")

            # Get the text content
            status = status_element.text.strip()
        except Exception as e:
            print(f"Error getting status for issue {issue_key}: {e}")

        # Get the priority
        priority = ""
        try:
            # Find the td with class "priority"
            priority_td = row.find_element(By.CSS_SELECTOR, "td.priority")

            # Find the img element inside the priority td
            priority_img = priority_td.find_element(By.CSS_SELECTOR, "img")

            # Get the alt attribute which contains the priority
            priority = priority_img.get_attribute("alt").strip()
        except Exception as e:
            print(f"Error getting priority for issue {issue_key}: {e}")

        # Get the Customer Object ID
        customer_object_id = ""
        try:
            # Find the td with class "customfield_14400"
            customfield_td = row.find_element(By.CSS_SELECTOR, "td.customfield_14400")

            # Get the text content directly from the td
            customer_object_id = customfield_td.text.strip()
        except Exception as e:
            print(f"Error getting Customer Object ID for issue {issue_key}: {e}")

        # Get the Assignee
        assignee = ""
        try:
            # Find the td with class "assignee"
            assignee_td = row.find_element(By.CSS_SELECTOR, "td.assignee")

            # Check if there's an <em> element (unassigned)
            try:
                em_element = assignee_td.find_element(By.CSS_SELECTOR, "em")
                assignee = em_element.text.strip()
            except:
                # If no <em>, try to find <a> element (user assigned)
                try:
                    a_element = assignee_td.find_element(
                        By.CSS_SELECTOR, "a.user-hover"
                    )
                    assignee = a_element.text.strip()
                except:
                    # If neither is found, get direct text from td
                    assignee = assignee_td.text.strip()
        except Exception as e:
            print(f"Error getting assignee for issue {issue_key}: {e}")

        # Get the Created date
        created_date = ""
        try:
            # Find the td with class "created"
            created_td = row.find_element(By.CSS_SELECTOR, "td.created")

            # Find the time element inside the created td
            time_element = created_td.find_element(By.CSS_SELECTOR, "time")

            # Get the datetime attribute
            iso_date = time_element.get_attribute("datetime")

            # Convert ISO string to Python datetime object without timezone info
            if iso_date:
                try:
                    # Parse the ISO date
                    dt_with_tz = parser.isoparse(iso_date)
                    # Convert to UTC and remove timezone information
                    created_date = dt_with_tz.astimezone(timezone.utc).replace(
                        tzinfo=None
                    )
                except Exception as e:
                    print(f"Error converting date format for issue {issue_key}: {e}")
                    created_date = (
                        iso_date  # Fallback to original format if error occurs
                    )
        except Exception as e:
            print(f"Error getting creation date for issue {issue_key}: {e}")

        # Get the Classification
        classification = ""
        try:
            # Find the td with class "customfield_15400"
            customfield_td = row.find_element(By.CSS_SELECTOR, "td.customfield_15400")

            # Get the text content directly from the td
            classification = customfield_td.text.strip()
        except Exception as e:
            print(f"Error getting Classification for issue {issue_key}: {e}")

        # Add data to lists if issue_key is valid
        if issue_key:
            issue_keys.append(issue_key)
            issue_types.append(issue_type)
            issue_links.append(issue_link)
            summaries.append(summary)
            statuses.append(status)
            priorities.append(priority)
            customer_object_ids.append(customer_object_id)
            assignees.append(assignee)
            created_dates.append(created_date)
            classifications.append(classification)

    # Create a DataFrame with all collected data
    df = pd.DataFrame(
        {
            "Issue Key": issue_keys,
            "Issue Type": issue_types,
            "Issue link": issue_links,
            "Summary": summaries,
            "Status": statuses,
            "Priority": priorities,
            "Customer Object ID": customer_object_ids,
            "Assignee": assignees,
            "Created": created_dates,
            "Classification": classifications,
        }
    )

    print(f"Found {len(issue_keys)} JIRA issues")

    return df


def adjust_column_widths(sheet, max_width=80):
    """
    Adjusts the width of each column in the Excel sheet based on the maximum width of the data and header values.

    Parameters:
    -----------
    sheet : openpyxl.worksheet.worksheet.Worksheet
        The worksheet where column widths need to be adjusted.

    max_width : int, optional (default=80)
        The maximum allowed width for any column. If the calculated width exceeds this value,
        the column width will be set to this maximum value.

    Returns:
    --------
    None
    """
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name

        # Calculate the width required by the header (considering formatting)
        header_length = len(str(col[0].value))
        adjusted_header_length = (
            header_length * 1.5
        )  # Factor to account for bold and larger font size

        # Compare the header length with the lengths of the data values
        for cell in col:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass

        # Use the greater of the header length or data length for column width
        max_length = max(max_length, adjusted_header_length)

        # Adjust the column width and apply the max_width limit
        adjusted_width = min(max_length + 2, max_width)
        sheet.column_dimensions[column].width = adjusted_width


def create_outlook_draft(
    excel_filename,
    recipient_list=None,
    cc_list=None,
    subject=None,
    body=None,
):
    """
    Creates an Outlook draft email with the Excel file attached.

    This function uses the win32com library to interact with Microsoft Outlook
    and create a draft email with the specified parameters. The email is saved
    in the Drafts folder for review before sending.

    Args:
        excel_filename (str): Full path to the Excel file to be attached
        recipient_list (str, optional): List of primary recipients separated by semicolons
        cc_list (str, optional): List of CC recipients separated by semicolons
        subject (str, optional): Email subject line
        body (str, optional): Email body content in HTML format

    Returns:
        bool: True if draft was successfully created, False otherwise
    """
    try:
        # Verify the file exists
        file_path = Path(excel_filename)
        if not file_path.exists():
            print(f"Error: File not found: {excel_filename}")
            return False

        print(f"Creating Outlook email draft...")

        # Initialize Outlook
        outlook = win32com.client.Dispatch("Outlook.Application")

        # Create a new message
        mail = outlook.CreateItem(0)  # 0 = olMailItem

        # Configure the message
        mail.To = recipient_list if recipient_list else ""

        # Add CC recipients if provided
        if cc_list:
            mail.CC = cc_list

        mail.Subject = subject
        mail.HTMLBody = body

        # Attach the file
        attachment = str(file_path.resolve())
        mail.Attachments.Add(attachment)

        # Save as draft
        mail.Save()

        print(
            f"Email draft created successfully. You can find it in your Outlook Drafts folder."
        )

        return True

    except Exception as e:
        print(f"Error creating email draft: {e}")
        return False


def generate_excel_report(excel_filename, dataframes_dict):
    """
    Generates an Excel file with multiple sheets from the provided dataframes.

    Applies formatting to each sheet:
    - Bold headers with blue background
    - Converts links to hyperlinks
    - Formats date columns appropriately
    - Adjusts column widths automatically
    - Adds filtering capability to all columns

    Args:
        excel_filename (str): Name of the Excel file to create
        dataframes_dict (dict): Dictionary with sheet names as keys and dataframes as values

    Returns:
        str: Path to the created Excel file
    """
    print(f"\nCreating Excel file with {len(dataframes_dict)} sheets: {excel_filename}")

    # Use ExcelWriter for more control over formatting
    with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
        # Create a sheet for each DataFrame
        for sheet_name, df in dataframes_dict.items():
            print(f"Creating sheet: {sheet_name} with {len(df)} issues")

            # Save DataFrame to its corresponding sheet
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Get the current worksheet object
            worksheet = writer.sheets[sheet_name]

            # Apply header formatting
            header_font = Font(bold=True)
            header_fill = PatternFill(
                start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
            )

            # Format the header row
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill

            # Convert "Issue link" column to hyperlinks and format date columns
            link_col_index = None
            date_col_index = None

            # Find relevant columns
            for i, header in enumerate(worksheet[1], 1):
                if header.value == "Issue link":
                    link_col_index = i
                elif header.value == "Created":
                    date_col_index = i

            # If link column found, convert to hyperlinks
            if link_col_index:
                link_col_letter = get_column_letter(link_col_index)

                # For each data row (starting from 2 since 1 is the header)
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet[f"{link_col_letter}{row}"]
                    if cell.value and isinstance(cell.value, str):
                        # Set hyperlink
                        cell.hyperlink = cell.value
                        # Apply hyperlink style (blue underlined)
                        cell.style = "Hyperlink"

            # If date column found, ensure proper formatting
            if date_col_index:
                date_col_letter = get_column_letter(date_col_index)

                # For each data row (starting from 2 since 1 is the header)
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet[f"{date_col_letter}{row}"]
                    if cell.value:
                        # Apply date and time format
                        cell.number_format = "yyyy-mm-dd hh:mm:ss"

            # Adjust column widths based on content
            adjust_column_widths(worksheet)

            # Apply filter to all columns
            worksheet.auto_filter.ref = worksheet.dimensions

    print(f"Excel file successfully created and formatted: {excel_filename}")
    return excel_filename


def generate_email_draft(excel_filename, dataframes_dict, current_time):
    """
    Creates an Outlook draft email with the Excel file attached and dynamic content.

    This function prepares an email draft by:
    - Loading recipients (To and CC) from environment variables
    - Generating a dynamic list of sheets with issue counts
    - Formatting the email body with placeholders
    - Creating the draft with the Excel file attached

    Args:
        excel_filename (str): Path to the Excel file to be attached
        dataframes_dict (dict): Dictionary containing all dataframes with issues data
        current_time (str): Current timestamp string for subject formatting

    Returns:
        bool: True if the draft was successfully created, False otherwise
    """
    print("\nPreparing email draft with JIRA issues report...")

    # Load recipients from environment variables
    recipients = load_mail_recipients("TO")
    cc_recipients = load_mail_recipients("CC")

    subject = os.getenv("MAIL_SUBJECT", "JIRA Issues Report - ")
    body_template = os.getenv("MAIL_BODY_TEMPLATE", "")

    # Generate list of sheets dynamically
    pestanas_html = "<ul>\n"
    for name, df in dataframes_dict.items():
        pestanas_html += f"<li>{name}: {len(df)} issues</li>\n"
    pestanas_html += "</ul>"

    # Replace placeholders with dynamic values
    body_formateado = body_template.format(
        FECHA=datetime.now().strftime("%d/%m/%Y %H:%M"),
        NUM_ISSUES=sum(len(df) for df in dataframes_dict.values()),
        NUM_PESTANAS=len(dataframes_dict),
        LISTA_PESTANAS=pestanas_html,
    )

    # Create the Outlook draft
    result = create_outlook_draft(
        excel_filename=excel_filename,
        recipient_list=recipients,
        cc_list=cc_recipients,
        subject=f"{subject}{current_time}",
        body=body_formateado,
    )

    return result


def remove_file(file_path):
    """
    Removes a file from the filesystem after it's no longer needed.

    Args:
        file_path (str): Path to the file to be removed

    Returns:
        bool: True if file was successfully removed, False otherwise
    """
    try:
        # Check if file exists before attempting to delete
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"File removed successfully: {file_path}")
            return True
        else:
            print(f"File not found: {file_path}")
            return False
    except Exception as e:
        print(f"Error removing file {file_path}: {e}")
        return False


def format_sheet_name(name):
    """
    Formats a sheet name to make it more readable:
    - Converts to lowercase
    - Replaces underscores with spaces
    - Capitalizes the first letter of each word

    Args:
        name (str): Original name to format

    Returns:
        str: Formatted sheet name
    """
    # Replace underscores with spaces and convert to lowercase
    formatted_name = name.lower().replace("_", " ")

    # Capitalize the first letter of each word (title case)
    formatted_name = formatted_name.title()

    return formatted_name


def main():
    """
    Main execution function that orchestrates the entire JIRA issues extraction workflow.

    This function performs the following steps:
    1. Loads environment variables from .env file
    2. Configures execution parameters (headless mode, timeouts, etc.)
    3. Loads JIRA filters from environment variables
    4. Initializes a Chrome WebDriver
    5. For each JIRA filter:
       - Builds a URL with the encoded filter
       - Navigates to the URL
       - Extracts JIRA issues into a DataFrame
       - Stores the DataFrame in a dictionary with the filter name as key
    6. Generates an Excel file with multiple sheets (one per filter)
    7. Creates an Outlook draft email with the Excel file attached
    8. Ensures proper cleanup of resources

    The function relies on environment variables defined in .env file:
    - HEADLESS_MODE: Whether to run Chrome in headless mode (True/False)
    - WAIT_TIME: Timeout in seconds for page loading
    - WAIT_ELEMENT: CSS selector of element to wait for
    - JIRA_URL_BASE: Base URL for JIRA
    - FILTER_PATTERN: Prefix for JIRA filter environment variables
    - Various JIRA filters (defined by FILTER_PATTERN)
    - Email configuration (recipients, subject, body template)

    No parameters or return values as this is the main orchestration function.
    """
    # Load environment variables from .env file
    load_dotenv()

    headless = os.getenv("HEADLESS_MODE", "False").lower() == "true"
    timeout = int(os.getenv("WAIT_TIME", "10"))
    wait_element = os.getenv("WAIT_ELEMENT", ".issue-table-wrapper")
    url_base = os.getenv("JIRA_URL_BASE")
    filter_pattern = os.getenv("FILTER_PATTERN", "JIRA_FILTER_")

    # Load all JIRA filters
    jira_filters = load_jira_filters(filter_pattern)
    print(f"Found {len(jira_filters)} JIRA filters to process")

    # Dictionary to store resulting dataframes
    dataframes_dict = {}

    # Create the driver
    driver = create_chrome_driver(headless=headless)

    try:
        # Iterate through each JIRA filter
        for filter_name, filter_value in jira_filters.items():
            print(f"\nProcessing filter: {filter_name}")

            # Extract sheet name (removing prefix)
            raw_sheet_name = filter_name.replace(filter_pattern, "")
            sheet_name = format_sheet_name(raw_sheet_name)

            # Encode JQL filter for safe URL use
            filter_encoded = quote(filter_value) if filter_value else ""

            # Build complete URL with encoded filter
            complete_url = f"{url_base}?jql={filter_encoded}"

            # Navigate to page with explicit wait for the issue table
            navigate_to_url(
                driver, complete_url, wait_element=wait_element, timeout=timeout
            )

            # Extract JIRA issues into a DataFrame
            jira_issues_df = extract_jira_issues(driver)

            # Store DataFrame in dictionary using sheet name as key
            dataframes_dict[sheet_name] = jira_issues_df

        # Generate filename with current date/time
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"jira_issues_{current_time}.xlsx"

        # Generate Excel file with all dataframes
        excel_filename = generate_excel_report(excel_filename, dataframes_dict)

        # Generate email draft with Excel attachment
        email_success = generate_email_draft(
            excel_filename, dataframes_dict, current_time
        )

        # Remove the Excel file after it's been attached to the email
        if email_success:
            print(
                "\nEmail draft created successfully. Removing temporary Excel file..."
            )
            remove_file(excel_filename)
        else:
            print("\nEmail draft creation failed. Keeping Excel file for reference.")

    finally:
        # Always close the driver to prevent resource leaks
        print("Closing browser...")
        driver.quit()


if __name__ == "__main__":
    main()
